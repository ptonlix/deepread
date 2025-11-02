"""
Document conversion entry points.

Each supported format is transformed into a list of `PageImage` instances whose
primary payload is a rendered canvas ready for OCR.

For PDF files, we preserve the original visual layout by converting pages directly
to images rather than extracting text and re-rendering.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
import tempfile
import os

from PIL import Image
from openpyxl import load_workbook
import pandas as pd
import imgkit

from pdf2image import convert_from_bytes
from html2image import Html2Image
from docx2pdf import convert

# DPI constant for image consistency
DPI = (300, 300)

SUPPORTED_FORMATS = {"pdf", "docx", "xlsx", "html", "htm"}


@dataclass(slots=True)
class PageImage:
    """Rendered representation of a logical page."""

    index: int
    image: Image.Image


def convert_document(payload: bytes, filename: str) -> list[PageImage]:
    """Dispatch conversion based on filename extension."""
    extension = filename.split(".")[-1].lower()
    ensure_supported(extension)

    if extension in {"html", "htm"}:
        html_pages = _convert_html(payload)
        return [
            PageImage(
                index=idx,
                image=image,
            )
            for idx, image in enumerate(html_pages)
            if image is not None
        ]
    elif extension == "docx":
        docx_pages = _convert_docx(payload)
        return [
            PageImage(
                index=idx,
                image=image,
            )
            for idx, image in enumerate(docx_pages)
            if image is not None
        ]
    elif extension == "xlsx":
        xlsx_pages = _convert_xlsx(payload)
        return [
            PageImage(
                index=idx,
                image=image,
            )
            for idx, image in enumerate(xlsx_pages)
            if image is not None
        ]
    elif extension == "pdf":
        pdf_pages = _convert_pdf(payload)
        return [
            PageImage(
                index=idx,
                image=image,
            )
            for idx, image in enumerate(pdf_pages)
            if image is not None
        ]
    else:  # pragma: no cover - safeguarded by ensure_supported
        raise ValueError(f"Unsupported document format: {extension}")


def ensure_supported(extension: str) -> None:
    if extension.lower() not in SUPPORTED_FORMATS:
        raise ValueError(f"Unsupported document format: {extension}")


def _convert_html(payload: bytes) -> list[Image.Image | None]:
    """Convert HTML to images using html2image."""
    try:
        html_content = payload.decode("utf-8")

        # Create Html2Image instance
        hti = Html2Image(size=(2480, 3508))  # A4 size at 300 DPI

        # Convert HTML to image
        # html2image saves to disk, so we need to handle that
        import tempfile
        import os

        with tempfile.TemporaryDirectory() as temp_dir:
            # Enhanced CSS for better rendering
            enhanced_css = """
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
                line-height: 1.6;
                color: #333;
                background-color: white;
            }
            h1, h2, h3, h4, h5, h6 {
                color: #2c3e50;
                margin-top: 1.5em;
                margin-bottom: 0.5em;
            }
            p {
                margin-bottom: 1em;
            }
            table {
                border-collapse: collapse;
                width: 100%;
                margin: 1em 0;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }
            th {
                background-color: #f2f2f2;
            }
            img {
                max-width: 100%;
                height: auto;
            }
            """

            # Generate image from HTML content
            image_paths = hti.screenshot(
                html_str=html_content,
                save_as="page.png",
                css_str=enhanced_css,
                size=(2480, 3508),
            )

            # Load the generated image
            if image_paths:
                image_path = os.path.join(temp_dir, "page.png")
                if os.path.exists(image_path):
                    image = Image.open(image_path).copy()
                    return [image]

        # If we reach here, conversion failed
        return [None]
    except Exception:
        # Return None image on any error
        return [None]


def _convert_docx(payload: bytes) -> list[Image.Image | None]:
    """Convert DOCX to images using docx2pdf and pdf2image."""
    result: list[Image.Image | None] = []

    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save DOCX file temporarily
            docx_path = os.path.join(temp_dir, "document.docx")
            with open(docx_path, "wb") as f:
                f.write(payload)

            # Convert DOCX to PDF
            pdf_path = os.path.join(temp_dir, "document.pdf")
            convert(docx_path, pdf_path)

            # Read the generated PDF and convert to images
            with open(pdf_path, "rb") as pdf_file:
                pdf_payload = pdf_file.read()
                images = convert_from_bytes(pdf_payload, dpi=300)

            # Create result with images
            for image in images:
                # Set DPI metadata for consistency
                image.info["dpi"] = DPI
                result.append(image)

            return result or [None]
    except Exception:
        # Return None image on any error
        return [None]


def _convert_xlsx(payload: bytes) -> list[Image.Image | None]:
    """Convert Excel to images using pandas and imgkit."""
    result: list[Image.Image | None] = []

    # Load workbook with openpyxl to get sheet names
    workbook = load_workbook(io.BytesIO(payload), read_only=True)

    with tempfile.TemporaryDirectory() as temp_dir:
        # Save Excel file temporarily
        excel_path = os.path.join(temp_dir, "workbook.xlsx")
        with open(excel_path, "wb") as f:
            f.write(payload)

        # Process each sheet
        for sheet_name in workbook.sheetnames:
            try:
                # Read sheet with pandas
                df = pd.read_excel(excel_path, sheet_name=sheet_name)

                # Skip empty sheets
                if df.empty:
                    continue

                # Convert DataFrame to HTML with enhanced styling
                html_content = df.to_html(
                    index=False,
                    table_id="excel-table",
                    classes="excel-table",
                    escape=False,
                )

                # Enhanced CSS for better rendering with Chinese font support
                css_content = """
                @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;700&display=swap');

                body {
                    font-family: 'Noto Sans SC', 'Microsoft YaHei', 'SimHei', 'Arial Unicode MS', sans-serif;
                    margin: 20px;
                    background-color: white;
                    font-size: 16px;
                    line-height: 1.4;
                }
                h2 {
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 20px;
                    color: #333;
                    border-bottom: 2px solid #4CAF50;
                    padding-bottom: 10px;
                }
                .excel-table {
                    border-collapse: collapse;
                    width: 100%;
                    margin: 20px 0;
                    font-size: 16px;
                    font-family: 'Noto Sans SC', 'Microsoft YaHei', 'SimHei', 'Arial Unicode MS', sans-serif;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                }
                .excel-table th,
                .excel-table td {
                    border: 1px solid #ddd;
                    padding: 12px 16px;
                    text-align: left;
                    vertical-align: middle;
                    word-wrap: break-word;
                    max-width: 200px;
                }
                .excel-table th {
                    background-color: #4CAF50;
                    color: white;
                    font-weight: bold;
                    font-size: 18px;
                    text-align: center;
                }
                .excel-table tr:nth-child(even) {
                    background-color: #f8f9fa;
                }
                .excel-table tr:hover {
                    background-color: #e8f5e8;
                }
                .excel-table td {
                    font-size: 16px;
                }
                """

                # Add sheet title
                full_html = f"""
                <html>
                <head>
                    <style>{css_content}</style>
                </head>
                <body>
                    <h2>{sheet_name}</h2>
                    {html_content}
                </body>
                </html>
                """

                # Generate image using imgkit with better options for Chinese
                options = {
                    "page-size": "A4",
                    "margin-top": "0.75in",
                    "margin-right": "0.75in",
                    "margin-bottom": "0.75in",
                    "margin-left": "0.75in",
                    "encoding": "UTF-8",
                    "no-outline": None,
                    "enable-local-file-access": None,
                    "width": 1200,
                    "height": 1600,
                    "quality": 100,
                    "format": "png",
                }

                # Convert HTML to image
                img_path = os.path.join(temp_dir, f"{sheet_name}.png")
                imgkit.from_string(full_html, img_path, options=options)

                # Load the generated image
                if os.path.exists(img_path):
                    image = Image.open(img_path).copy()
                    # Set DPI metadata for consistency
                    image.info["dpi"] = DPI
                    result.append(image)
                else:
                    # Fallback: no image available
                    result.append(None)

            except Exception:
                # Fallback: no image available for this sheet
                result.append(None)

    return result or [None]


def _convert_pdf(payload: bytes) -> list[Image.Image | None]:
    """Convert PDF to images, preserving original visual layout."""
    # Convert PDF pages directly to images
    images = convert_from_bytes(payload, dpi=300)
    result: list[Image.Image | None] = []

    for image in images:
        # Set DPI metadata for consistency
        image.info["dpi"] = DPI
        result.append(image)

    return result
